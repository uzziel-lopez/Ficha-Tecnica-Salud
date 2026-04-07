#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para integrar datos de recursos del IFU en las fichas técnicas
"""

import json
import os
import sys
import csv
import pymssql
from pathlib import Path
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[2]
DATA_INPUT_DIR = BASE_DIR / "data" / "input"
DATA_OUTPUT_DIR = BASE_DIR / "data" / "output"


def read_env(name, default=""):
    return os.getenv(name, default).strip()


SIAIS_DB_CONFIG = {
    "user": read_env("FT_SIAIS_DB_USER"),
    "password": read_env("FT_SIAIS_DB_PASSWORD"),
    "database": read_env("FT_SIAIS_DB_NAME", "DEMO_DB"),
    "port": int(read_env("FT_SIAIS_DB_PORT", "15433")),
}


def get_siais_servers():
    raw_servers = read_env("FT_SIAIS_SERVER_LIST", "")
    if raw_servers:
        return [item.strip() for item in raw_servers.split(",") if item.strip()]
    return [
        "192.0.2.10",
        "192.0.2.11",
    ]

# Configurar la salida UTF-8 para Windows
if sys.platform == "win32":
    import codecs

    sys.stdout = codecs.getwriter("utf-8")(sys.stdout.buffer, "strict")
    sys.stderr = codecs.getwriter("utf-8")(sys.stderr.buffer, "strict")


def cargar_datos_ifu(archivo_ifu=str(DATA_OUTPUT_DIR / "claves_ifu_extraidas.json")):
    """Carga los datos extraídos del IFU"""
    if not os.path.exists(archivo_ifu):
        print(f" No se encontró el archivo: {archivo_ifu}")
        print(" Archivos JSON disponibles:")
        for archivo in os.listdir(DATA_OUTPUT_DIR):
            if archivo.startswith("claves_ifu_extraidas") and archivo.endswith(".json"):
                print(f"   - {archivo}")
        return None

    print(f" Cargando datos del IFU: {archivo_ifu}")

    with open(archivo_ifu, "r", encoding="utf-8") as f:
        datos_ifu = json.load(f)

    unidades_con_datos = datos_ifu.get("unidades", [])
    print(f" Cargados datos de {len(unidades_con_datos)} unidades")

    return unidades_con_datos


def calculate_business_days_for_month(year, month):
    """
    Calcula días hábiles del 26 del mes anterior al 25 del mes actual
    Excluye fines de semana y días festivos oficiales
    """
    try:
        # Días festivos por mes (cantidad a restar)
        DIAS_FESTIVOS_POR_MES = {
            1: 1,  # Enero: 1 día festivo (1 de enero)
            2: 1,  # Febrero: 1 día festivo (5 de febrero)
            3: 1,  # Marzo: 1 día festivo (21 de marzo)
            4: 2,  # Abril: 2 días festivos
            5: 2,  # Mayo: 2 días festivos (1 y 5 de mayo)
            6: 0,  # Junio: sin festivos
            7: 0,  # Julio: sin festivos
            8: 0,  # Agosto: sin festivos
            9: 2,  # Septiembre: 2 días festivos (16 de septiembre)
            10: 0,  # Octubre: sin festivos
            11: 1,  # Noviembre: 1 día festivo (20 de noviembre)
            12: 1,  # Diciembre: 1 día festivo (25 de diciembre)
        }

        # Determinar el mes anterior
        if month == 1:
            mes_anterior = 12
            year_anterior = year - 1
        else:
            mes_anterior = month - 1
            year_anterior = year

        # Fecha de inicio: 26 del mes anterior
        start_date = datetime(year_anterior, mes_anterior, 26)

        # Fecha de fin: 25 del mes actual
        end_date = datetime(year, month, 25)

        # Contar días hábiles (lunes a viernes)
        business_days = 0
        current_date = start_date

        while current_date <= end_date:
            # Solo contar si es día de semana (0=Lunes, 4=Viernes)
            if current_date.weekday() < 5:
                business_days += 1
            current_date += timedelta(days=1)

        # Restar días festivos del mes
        dias_festivos = DIAS_FESTIVOS_POR_MES.get(month, 0)
        business_days_finales = business_days - dias_festivos

        # Asegurar que no sea negativo
        business_days_finales = max(business_days_finales, 1)

        return business_days_finales

    except Exception as e:
        print(f"ERROR en calculate_business_days_for_month: {e}")
        return 22  # Promedio de días hábiles por mes


def cargar_fichas_originales(
    archivo=str(DATA_OUTPUT_DIR / "fichas_completas_tabasco.json"),
):
    """Carga las fichas técnicas originales"""
    print(f" Cargando fichas originales: {archivo}")

    with open(archivo, "r", encoding="utf-8") as f:
        fichas = json.load(f)

    unidades = fichas.get("unidades", [])
    print(f" Cargadas {len(unidades)} fichas originales")

    return fichas


def cargar_productividad_csv(
    archivo=str(DATA_INPUT_DIR / "PRODUCTIVIDAD_TRATAMIENTO" / "Productividad.csv"),
):
    """Carga el CSV de productividad de servicios auxiliares"""
    if not os.path.exists(archivo):
        print(f"  No se encontró el archivo: {archivo}")
        return {}

    print(f" Cargando datos de productividad: {archivo}")

    # Diccionario para almacenar datos por unidad
    datos_productividad = {}

    try:
        with open(archivo, "r", encoding="utf-16-le") as f:
            reader = csv.DictReader(f, delimiter="\t")

            for row in reader:
                unidad = row.get("Unidad", "").strip()
                descripcion = row.get("DESCRIPCION", "").strip()
                año = row.get("Año", "").strip()
                mes = row.get("Mes", "").strip()
                total_estudios = row.get("Total de estudios o sesiones", "0").strip()

                # Filtrar filas de totales y validar datos
                if not unidad or año == "Total" or mes == "Total":
                    continue

                # Crear estructura por unidad si no existe
                if unidad not in datos_productividad:
                    datos_productividad[unidad] = {}

                # Crear estructura por servicio si no existe
                if descripcion not in datos_productividad[unidad]:
                    datos_productividad[unidad][descripcion] = {}

                # Guardar valor por mes
                try:
                    datos_productividad[unidad][descripcion][mes] = int(
                        float(total_estudios)
                    )
                except ValueError:
                    datos_productividad[unidad][descripcion][mes] = 0

        print(
            f" Datos de productividad cargados: {len(datos_productividad)} unidades"
        )
        return datos_productividad

    except Exception as e:
        print(f" Error cargando productividad CSV: {e}")
        return {}


def procesar_productividad_unidad(unidad_clues, datos_productividad):
    """Procesa los datos de productividad para una unidad específica"""

    # Obtener datos de esta unidad
    datos_unidad = datos_productividad.get(unidad_clues, {})

    # Extraer TODOS los servicios/descripciones disponibles para esta unidad
    # en lugar de una lista fija
    servicios_mostrar = sorted(datos_unidad.keys())

    # Meses en el orden correcto con números de mes
    meses_info = [
        ("Enero", 1),
        ("Febrero", 2),
        ("Marzo", 3),
        ("Abril", 4),
        ("Mayo", 5),
        ("Junio", 6),
        ("Julio", 7),
        ("Agosto", 8),
        ("Septiembre", 9),
        ("Octubre", 10),
        ("Noviembre", 11),
        ("Diciembre", 12),
    ]

    # Año actual
    año_actual = datetime.now().year

    # Estructura final
    productividad_servicios_auxiliares = {}

    for servicio in servicios_mostrar:
        datos_servicio = datos_unidad.get(servicio, {})

        productividad_servicios_auxiliares[servicio] = {
            "meses": {},
            "dia_tipico": {},
            "total_acumulado": 0,
            "promedio_dia_tipico": 0,
        }

        total = 0
        total_dias_tipicos = 0
        meses_con_datos = 0

        for mes_nombre, mes_num in meses_info:
            valor = datos_servicio.get(mes_nombre, 0)
            productividad_servicios_auxiliares[servicio]["meses"][mes_nombre] = valor
            total += valor

            # Calcular día típico
            dias_habiles = calculate_business_days_for_month(año_actual, mes_num)
            dia_tipico = (
                round(valor / dias_habiles, 2) if dias_habiles > 0 and valor > 0 else 0
            )
            productividad_servicios_auxiliares[servicio]["dia_tipico"][mes_nombre] = (
                dia_tipico
            )

            if valor > 0:
                total_dias_tipicos += dia_tipico
                meses_con_datos += 1

        productividad_servicios_auxiliares[servicio]["total_acumulado"] = total
        productividad_servicios_auxiliares[servicio]["promedio_dia_tipico"] = (
            round(total_dias_tipicos / meses_con_datos, 2) if meses_con_datos > 0 else 0
        )

    return productividad_servicios_auxiliares


def cargar_catalogo_descripciones(
    archivo=str(DATA_INPUT_DIR / "Diagnosticos y Procedimientos.xlsx"),
):
    """Carga catálogo de descripciones CIE10 y CIE9"""
    if not os.path.exists(archivo):
        print(f"  No se encontró el archivo: {archivo}")
        return {}, {}

    print(f" Cargando catálogo de descripciones: {archivo}")

    try:
        # Hoja 1: CIE10 (Diagnósticos)
        workbook_cie10 = load_workbook(archivo, read_only=True, data_only=True)
        sheet_cie10 = workbook_cie10.worksheets[0]  # Primera hoja

        catalogo_cie10 = {}
        for row in sheet_cie10.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Procedimiento | Descripcion
                codigo = str(row[0]).strip()
                descripcion = str(row[1]).strip()
                catalogo_cie10[codigo] = descripcion

        workbook_cie10.close()

        # Hoja 2: CIE9 (Procedimientos Quirúrgicos)
        workbook_cie9 = load_workbook(archivo, read_only=True, data_only=True)
        sheet_cie9 = workbook_cie9.worksheets[1]  # Segunda hoja

        catalogo_cie9 = {}
        for row in sheet_cie9.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:  # Procedimiento | Descripcion
                codigo = str(row[0]).strip()
                descripcion = str(row[1]).strip()
                catalogo_cie9[codigo] = descripcion

        workbook_cie9.close()

        print(f"    CIE10: {len(catalogo_cie10)} diagnósticos cargados")
        print(f"    CIE9: {len(catalogo_cie9)} procedimientos cargados")

        return catalogo_cie10, catalogo_cie9

    except Exception as e:
        print(f" Error cargando catálogo: {e}")
        return {}, {}


def cargar_egresos_pacientes_hospital(archivo_excel):
    """
    Carga el Excel de Egresos de Pacientes y cuenta diagnósticos por hospital
    """
    if not os.path.exists(archivo_excel):
        print(f"  No se encontró el archivo: {archivo_excel}")
        return {}

    print(f" Cargando egresos de pacientes: {archivo_excel}")

    try:
        workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
        sheet = workbook.active

        # Leer encabezados
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # Verificar columnas requeridas
        if "cvePresupuestal" not in headers or "DiagPrincipalEgreso" not in headers:
            print(f" Faltan columnas requeridas en Egresos")
            return {}

        col_cve_presup = headers["cvePresupuestal"]
        col_diag = headers["DiagPrincipalEgreso"]

        # Diccionario: {clave_presup: {diagnostico: conteo}}
        datos_egresos = {}
        filas_procesadas = 0

        for row in sheet.iter_rows(min_row=2, values_only=False):
            try:
                cve_presup = (
                    str(row[col_cve_presup - 1].value).strip()
                    if row[col_cve_presup - 1].value
                    else None
                )
                diagnostico = (
                    str(row[col_diag - 1].value).strip()
                    if row[col_diag - 1].value
                    else None
                )

                if not cve_presup or not diagnostico:
                    continue

                if cve_presup not in datos_egresos:
                    datos_egresos[cve_presup] = {}

                if diagnostico not in datos_egresos[cve_presup]:
                    datos_egresos[cve_presup][diagnostico] = 0

                datos_egresos[cve_presup][diagnostico] += 1
                filas_procesadas += 1

            except Exception as e:
                continue

        workbook.close()

        print(f"    Procesadas {filas_procesadas} filas de egresos")
        print(f"    Hospitales con datos: {len(datos_egresos)}")

        return datos_egresos

    except Exception as e:
        print(f" Error procesando Egresos: {e}")
        return {}


def cargar_intervenciones_quirurgicas_hospital(archivo_excel):
    """
    Carga el Excel de Intervenciones Quirúrgicas y cuenta procedimientos por hospital
    """
    if not os.path.exists(archivo_excel):
        print(f"  No se encontró el archivo: {archivo_excel}")
        return {}

    print(f" Cargando intervenciones quirúrgicas: {archivo_excel}")

    try:
        workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
        sheet = workbook.active

        # Leer encabezados
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # Verificar columnas requeridas
        if "cvePresupuestal" not in headers or "PROCQ" not in headers:
            print(f" Faltan columnas requeridas en Intervenciones")
            return {}

        col_cve_presup = headers["cvePresupuestal"]
        col_procq = headers["PROCQ"]

        # Diccionario: {clave_presup: {procedimiento: conteo}}
        datos_intervenciones = {}
        filas_procesadas = 0

        for row in sheet.iter_rows(min_row=2, values_only=False):
            try:
                cve_presup = (
                    str(row[col_cve_presup - 1].value).strip()
                    if row[col_cve_presup - 1].value
                    else None
                )
                procedimiento = (
                    str(row[col_procq - 1].value).strip()
                    if row[col_procq - 1].value
                    else None
                )

                if not cve_presup or not procedimiento:
                    continue

                if cve_presup not in datos_intervenciones:
                    datos_intervenciones[cve_presup] = {}

                if procedimiento not in datos_intervenciones[cve_presup]:
                    datos_intervenciones[cve_presup][procedimiento] = 0

                datos_intervenciones[cve_presup][procedimiento] += 1
                filas_procesadas += 1

            except Exception as e:
                continue

        workbook.close()

        print(f"    Procesadas {filas_procesadas} filas de intervenciones")
        print(f"    Hospitales con datos: {len(datos_intervenciones)}")

        return datos_intervenciones

    except Exception as e:
        print(f" Error procesando Intervenciones: {e}")
        return {}


def procesar_motivos_hospital(
    clave_presupuestal,
    datos_egresos,
    datos_intervenciones,
    catalogo_cie10,
    catalogo_cie9,
):
    """
    Procesa motivos de atención para un hospital específico
    Retorna Top 25 diagnósticos de egreso y Top 25 procedimientos quirúrgicos
    """
    # Obtener datos de este hospital
    egresos_hospital = datos_egresos.get(clave_presupuestal, {})
    intervenciones_hospital = datos_intervenciones.get(clave_presupuestal, {})

    # Top 25 diagnósticos de egreso
    diagnosticos_ordenados = sorted(
        egresos_hospital.items(), key=lambda x: x[1], reverse=True
    )[:25]
    diagnosticos_egreso = []
    for codigo, total in diagnosticos_ordenados:
        diagnosticos_egreso.append(
            {
                "codigo": codigo,
                "descripcion": catalogo_cie10.get(codigo, "Sin descripción"),
                "total": total,
            }
        )

    # Top 25 procedimientos quirúrgicos
    procedimientos_ordenados = sorted(
        intervenciones_hospital.items(), key=lambda x: x[1], reverse=True
    )[:25]
    procedimientos_quirurgicos = []
    for codigo, total in procedimientos_ordenados:
        procedimientos_quirurgicos.append(
            {
                "codigo": codigo,
                "descripcion": catalogo_cie9.get(codigo, "Sin descripción"),
                "total": total,
            }
        )

    return {
        "diagnosticos_egreso": diagnosticos_egreso,
        "procedimientos_quirurgicos": procedimientos_quirurgicos,
    }


def procesar_consulta_externa_motivos_hospital(
    archivo_excel, clave_presupuestal, catalogo_cie10
):
    """
    Procesa el Excel de Consulta Externa Diaria para generar 2 tablas de motivos:
    1. MOTIVOS CONSULTAS DE ESPECIALIDADES (todas especialidades EXCEPTO paramédicos)
    2. MOTIVO URGENCIAS PRIMER CONTACTO (solo especialidades 5001, A600)
    """
    if not os.path.exists(archivo_excel):
        print(f"        Archivo no encontrado: {archivo_excel}")
        return None

    try:
        print(f"       Cargando {os.path.basename(archivo_excel)}...")
        wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
        ws = wb.active

        # Leer encabezados
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[cell.value.strip()] = col_idx

        # Verificar columnas necesarias
        columnas_requeridas = ["CVE_PRESUPUESTAL", "ESPECIALIDAD", "DIAG_PRINCIPAL"]
        for col in columnas_requeridas:
            if col not in headers:
                print(f"       Columna '{col}' no encontrada")
                wb.close()
                return None

        # Índices de columnas
        idx_cve = headers["CVE_PRESUPUESTAL"]
        idx_esp = headers["ESPECIALIDAD"]
        idx_diag = headers["DIAG_PRINCIPAL"]

        # Paramédicos a excluir
        paramedicos_excluir = {"39EG", "6300", "6400", "6600", "6900"}

        # Urgencias a incluir
        urgencias_incluir = {"5001", "A600"}

        # Contadores
        diagnosticos_especialidades = {}  # Todas especialidades excepto paramédicos
        diagnosticos_urgencias = {}  # Solo urgencias (5001, A600)

        total_rows = 0
        filas_procesadas = 0

        # Procesar filas
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_rows += 1

            # Obtener valores
            cve = str(row[idx_cve - 1]).strip() if row[idx_cve - 1] else ""
            especialidad = (
                str(row[idx_esp - 1]).strip().upper() if row[idx_esp - 1] else ""
            )
            diagnostico = str(row[idx_diag - 1]).strip() if row[idx_diag - 1] else ""

            # Filtrar por clave presupuestal
            if cve != str(clave_presupuestal):
                continue

            # Validar que tenga diagnóstico
            if not diagnostico or diagnostico == "None":
                continue

            filas_procesadas += 1

            # TABLA 1: Especialidades (excluyendo paramédicos)
            if especialidad and especialidad not in paramedicos_excluir:
                if diagnostico not in diagnosticos_especialidades:
                    diagnosticos_especialidades[diagnostico] = 0
                diagnosticos_especialidades[diagnostico] += 1

            # TABLA 2: Urgencias (solo 5001, A600)
            if especialidad in urgencias_incluir:
                if diagnostico not in diagnosticos_urgencias:
                    diagnosticos_urgencias[diagnostico] = 0
                diagnosticos_urgencias[diagnostico] += 1

        wb.close()

        print(
            f"       Procesadas {filas_procesadas:,} filas de {total_rows:,} totales"
        )
        print(
            f"       Diagnósticos únicos - Especialidades: {len(diagnosticos_especialidades)}, Urgencias: {len(diagnosticos_urgencias)}"
        )

        # Ordenar y obtener Top 25
        top_especialidades = sorted(
            diagnosticos_especialidades.items(), key=lambda x: x[1], reverse=True
        )[:25]
        top_urgencias = sorted(
            diagnosticos_urgencias.items(), key=lambda x: x[1], reverse=True
        )[:25]

        # Agregar descripciones
        especialidades_con_desc = []
        for codigo, total in top_especialidades:
            especialidades_con_desc.append(
                {
                    "clave": codigo,
                    "descripcion": catalogo_cie10.get(codigo, "Sin descripción"),
                    "total": total,
                }
            )

        urgencias_con_desc = []
        for codigo, total in top_urgencias:
            urgencias_con_desc.append(
                {
                    "clave": codigo,
                    "descripcion": catalogo_cie10.get(codigo, "Sin descripción"),
                    "total": total,
                }
            )

        return {
            "especialidades": especialidades_con_desc,
            "urgencias": urgencias_con_desc,
        }

    except Exception as e:
        print(f"       Error procesando consulta externa: {e}")
        return None


def cargar_consulta_externa_hgz(archivo_excel):
    """
    Carga el Excel de Consulta Externa Diaria para hospitales HGZ
    Procesa datos por mes según las reglas especificadas
    """
    if not os.path.exists(archivo_excel):
        print(f"  No se encontró el archivo: {archivo_excel}")
        return {}

    print(f" Cargando datos de Consulta Externa HGZ: {archivo_excel}")

    # Códigos de especialidad válidos
    ESPECIALIDAD_CODES = [
        "10",
        "11",
        "14",
        "15",
        "16",
        "18",
        "19",
        "20",
        "21",
        "22",
        "23",
        "25",
        "26",
        "27",
        "28",
        "30",
        "31",
        "32",
        "34",
        "35",
        "36",
        "38",
        "39",
        "41",
        "42",
        "43",
        "44",
        "45",
        "46",
        "51",
        "65",
        "6800",
    ]

    # Especialidades que NO cuentan para "Consultas de Especialidades"
    ESPECIALIDADES_EXCLUIDAS = ["39EG", "6300", "6400", "6600", "6900"]

    # Claves presupuestales de hospitales
    HOSPITALES_CLAVES = [
        "280132012151",  # HGZ 46 Villahermosa
        "280202012151",  # HGZ 2 Cárdenas
        "280208012151",  # HGZ 2A Cárdenas
        "280803052151",  # HGSMF 4 Tenosique
    ]

    try:
        workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
        sheet = workbook.active

        # Diccionario para almacenar datos por clave presupuestal y mes
        datos_hospitales = {}

        # Leer encabezados para encontrar las columnas
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip().upper()] = col_idx

        # Verificar que existan las columnas necesarias
        columnas_requeridas = [
            "MES_ORIGEN",
            "CVE_PRESUPUESTAL",
            "ESPECIALIDAD",
            "PRIMERA_VEZ",
        ]
        columnas_faltantes = [col for col in columnas_requeridas if col not in headers]

        if columnas_faltantes:
            print(f" Faltan columnas en el Excel: {columnas_faltantes}")
            return {}

        # Índices de columnas (1-based)
        col_mes = headers["MES_ORIGEN"]  # Columna AR
        col_cve_presupuestal = headers["CVE_PRESUPUESTAL"]  # Columna B
        col_especialidad = headers["ESPECIALIDAD"]  # Columna D
        col_primera_vez = headers["PRIMERA_VEZ"]  # Columna J

        print(
            f"    Columnas encontradas: MES={col_mes}, CVE_PRESUP={col_cve_presupuestal}, ESP={col_especialidad}, 1VEZ={col_primera_vez}"
        )

        # Procesar cada fila (empezar desde fila 2, saltar encabezados)
        filas_procesadas = 0
        for row in sheet.iter_rows(min_row=2, values_only=False):
            try:
                # Obtener valores
                mes_origen = (
                    str(row[col_mes - 1].value).strip().upper()
                    if row[col_mes - 1].value
                    else None
                )
                cve_presup = (
                    str(row[col_cve_presupuestal - 1].value).strip()
                    if row[col_cve_presupuestal - 1].value
                    else None
                )
                especialidad = (
                    str(row[col_especialidad - 1].value).strip()
                    if row[col_especialidad - 1].value
                    else None
                )
                primera_vez = row[col_primera_vez - 1].value

                # Filtrar solo hospitales válidos
                if not cve_presup or cve_presup not in HOSPITALES_CLAVES:
                    continue

                # Filtrar solo especialidades válidas
                if not especialidad or (
                    especialidad not in ESPECIALIDAD_CODES
                    and especialidad not in ESPECIALIDADES_EXCLUIDAS
                    and especialidad not in ["5001", "A600"]
                ):
                    continue

                # Normalizar mes
                if not mes_origen or mes_origen == "TOTAL":
                    continue

                # Inicializar estructura si no existe
                if cve_presup not in datos_hospitales:
                    datos_hospitales[cve_presup] = {}

                if mes_origen not in datos_hospitales[cve_presup]:
                    datos_hospitales[cve_presup][mes_origen] = {
                        "total_consultas": 0,
                        "consultas_especialidades": 0,
                        "consultas_primera_vez": 0,
                        "consultas_subsecuentes": 0,
                        "consultas_urgencias": 0,
                        "consultas_tococirrugia": 0,
                        "servicios_profesionales": 0,
                    }

                # Incrementar contadores según reglas
                mes_data = datos_hospitales[cve_presup][mes_origen]

                # Total de Consultas de la Unidad: TODAS las especialidades
                mes_data["total_consultas"] += 1

                # Consultas de Primera Vez: todas las especialidades con PRIMERA_VEZ = 1
                if primera_vez == 1 or primera_vez == "1":
                    mes_data["consultas_primera_vez"] += 1
                else:
                    # Consultas Subsecuentes: total - primera vez
                    mes_data["consultas_subsecuentes"] += 1

                # Consultas de Especialidades: TODAS excepto las excluidas
                if (
                    especialidad not in ESPECIALIDADES_EXCLUIDAS
                    and especialidad not in ["5001", "A600"]
                ):
                    mes_data["consultas_especialidades"] += 1

                # Consultas de Urgencias: solo especialidad 5001
                if especialidad == "5001":
                    mes_data["consultas_urgencias"] += 1

                # Consultas de Tococirugía: solo especialidad A600
                if especialidad == "A600":
                    mes_data["consultas_tococirrugia"] += 1

                # Servicios Profesionales: solo especialidades excluidas
                if especialidad in ESPECIALIDADES_EXCLUIDAS:
                    mes_data["servicios_profesionales"] += 1

                filas_procesadas += 1

            except Exception as e:
                # Error en fila específica, continuar
                continue

        workbook.close()

        print(f"    Procesadas {filas_procesadas} filas de consulta externa")
        print(f"    Hospitales con datos: {len(datos_hospitales)}")

        # DEBUG: Mostrar resumen por hospital
        for clave_presup, datos_hospital in datos_hospitales.items():
            total_consultas = sum(
                mes_data["total_consultas"] for mes_data in datos_hospital.values()
            )
            print(
                f"       {clave_presup}: {len(datos_hospital)} meses, {total_consultas} consultas totales"
            )

        return datos_hospitales

    except Exception as e:
        print(f" Error procesando Excel de Consulta Externa: {e}")
        return {}


def cargar_hospitalizacion_parteii(archivo_excel):
    """
    Carga el Excel PARTEII_2025.xlsx con datos de hospitalización
    Procesa datos por División y mes para HGZ y HGSZMF
    """
    if not os.path.exists(archivo_excel):
        print(f"  No se encontró el archivo: {archivo_excel}")
        return {}

    print(f" Cargando datos de Hospitalización PARTE II: {archivo_excel}")

    try:
        workbook = load_workbook(archivo_excel, read_only=True, data_only=True)
        sheet = workbook.active

        # Diccionario para almacenar datos por clave presupuestal, división y mes
        datos_hospitalizacion = {}

        # Leer encabezados
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # Verificar columnas requeridas
        columnas_requeridas = [
            "NombreDelegacionUMAE",
            "ClavePresupuestal",
            "Año",
            "mes",
            "Division",
            "Cve Especialidad",
            "Especialidad",
            "Camas",
            "Total Egreso",
            "Dias Paciente",
            "Porc Ocupacion",
            "Prom Estancia",
            "ind_rotacion",
            "ind_sustitucion",
            "Total Iq",
            "Egr Defuncion",
        ]

        columnas_faltantes = [col for col in columnas_requeridas if col not in headers]
        if columnas_faltantes:
            print(f" Faltan columnas en PARTEII: {columnas_faltantes}")
            return {}

        # Índices de columnas (1-based)
        col_delegacion = headers["NombreDelegacionUMAE"]
        col_cve_presup = headers["ClavePresupuestal"]
        col_año = headers["Año"]
        col_mes = headers["mes"]
        col_division = headers["Division"]
        col_cve_especialidad = headers["Cve Especialidad"]
        col_especialidad = headers["Especialidad"]
        col_camas = headers["Camas"]
        col_total_egreso = headers["Total Egreso"]
        col_dias_paciente = headers["Dias Paciente"]
        col_porc_ocupacion = headers["Porc Ocupacion"]
        col_prom_estancia = headers["Prom Estancia"]
        col_ind_rotacion = headers["ind_rotacion"]
        col_ind_sustitucion = headers["ind_sustitucion"]
        col_total_iq = headers["Total Iq"]
        col_egr_defuncion = headers["Egr Defuncion"]

        print(f"    Columnas encontradas correctamente")

        # Mapeo de números de mes a nombres
        meses_map = {
            1: "ENERO",
            2: "FEBRERO",
            3: "MARZO",
            4: "ABRIL",
            5: "MAYO",
            6: "JUNIO",
            7: "JULIO",
            8: "AGOSTO",
            9: "SEPTIEMBRE",
            10: "OCTUBRE",
            11: "NOVIEMBRE",
            12: "DICIEMBRE",
        }

        # Procesar cada fila
        filas_procesadas = 0
        año_actual = datetime.now().year

        for row in sheet.iter_rows(min_row=2, values_only=False):
            try:
                # Obtener valores
                delegacion = (
                    str(row[col_delegacion - 1].value).strip()
                    if row[col_delegacion - 1].value
                    else None
                )
                cve_presup = (
                    str(row[col_cve_presup - 1].value).strip()
                    if row[col_cve_presup - 1].value
                    else None
                )
                año = row[col_año - 1].value
                mes_num = row[col_mes - 1].value
                division = (
                    str(row[col_division - 1].value).strip()
                    if row[col_division - 1].value
                    else None
                )
                cve_especialidad = (
                    str(row[col_cve_especialidad - 1].value).strip()
                    if row[col_cve_especialidad - 1].value
                    else None
                )

                # Filtrar solo Tabasco y año actual
                if not delegacion or delegacion != "Tabasco":
                    continue
                if not año or int(año) != año_actual:
                    continue
                if not cve_presup or not division or not mes_num:
                    continue

                # Normalizar división
                division = division.strip()
                mes_nombre = meses_map.get(int(mes_num), None)
                if not mes_nombre:
                    continue

                # Inicializar estructura
                if cve_presup not in datos_hospitalizacion:
                    datos_hospitalizacion[cve_presup] = {}

                if mes_nombre not in datos_hospitalizacion[cve_presup]:
                    datos_hospitalizacion[cve_presup][mes_nombre] = {
                        "Cirugía": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                        "Gineco-Obstetricia": {
                            "egresos": 0,
                            "dias_paciente": 0,
                            "camas": 0,
                        },
                        "Medicina": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                        "Pediatría": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                        "Registros Especiales": {
                            "cirugias_ambulatorias": 0,
                            "egresos_urgencias_observacion": 0,
                            "uci": 0,
                            "uci_neonato": 0,
                            "urg_reanimacion": 0,
                            "urg_observacion_intermedia": 0,
                            "urg_observacion_pediatria": 0,
                            "urgencias_tococirugia": 0,
                            "ecam": 0,
                        },
                    }

                # Acumular valores según división
                mes_data = datos_hospitalizacion[cve_presup][mes_nombre]

                # Acumular valores según división
                total_egreso = (
                    float(row[col_total_egreso - 1].value)
                    if row[col_total_egreso - 1].value
                    else 0
                )

                if division in [
                    "Cirugía",
                    "Gineco-Obstetricia",
                    "Medicina",
                    "Pediatría",
                ]:
                    # IMPORTANTE: En el Excel hay filas de especialidades individuales Y filas TOTALES por división
                    # Para evitar sumar dos veces, tomamos el MÁXIMO valor (la fila TOTAL tendrá el mayor valor)
                    dias_cama = (
                        float(row[col_camas - 1].value)
                        if row[col_camas - 1].value
                        else 0
                    )
                    dias_paciente = (
                        float(row[col_dias_paciente - 1].value)
                        if row[col_dias_paciente - 1].value
                        else 0
                    )

                    # Tomar el MÁXIMO (no sumar) para usar la fila TOTAL de cada división
                    mes_data[division]["egresos"] = max(
                        mes_data[division]["egresos"], total_egreso
                    )
                    mes_data[division]["dias_paciente"] = max(
                        mes_data[division]["dias_paciente"], dias_paciente
                    )
                    mes_data[division]["camas"] = max(
                        mes_data[division]["camas"], dias_cama
                    )

                elif division == "Registros especiales":
                    # Para Registros especiales, usar la columna Cve Especialidad para identificar cada tipo
                    # y tomar el valor de Total Egreso (columna U)
                    # Mapeo de códigos de especialidad a campos de Registros Especiales
                    if cve_especialidad == "0100":  # Cirugía Ambulatoria
                        mes_data["Registros Especiales"]["cirugias_ambulatorias"] = max(
                            mes_data["Registros Especiales"]["cirugias_ambulatorias"],
                            total_egreso,
                        )
                    elif cve_especialidad == "4800":  # Unidad de Cuidados Intensivos
                        mes_data["Registros Especiales"]["uci"] = max(
                            mes_data["Registros Especiales"]["uci"], total_egreso
                        )
                    elif cve_especialidad == "4900":  # UCI Neonatos
                        mes_data["Registros Especiales"]["uci_neonato"] = max(
                            mes_data["Registros Especiales"]["uci_neonato"],
                            total_egreso,
                        )
                    elif cve_especialidad == "5002":  # Urgencias Área Reanimación
                        mes_data["Registros Especiales"]["urg_reanimacion"] = max(
                            mes_data["Registros Especiales"]["urg_reanimacion"],
                            total_egreso,
                        )
                    elif cve_especialidad == "5003":  # Urgencias Observación Intermedia
                        mes_data["Registros Especiales"][
                            "urg_observacion_intermedia"
                        ] = max(
                            mes_data["Registros Especiales"][
                                "urg_observacion_intermedia"
                            ],
                            total_egreso,
                        )
                    elif cve_especialidad == "5004":  # Urgencias Observación
                        mes_data["Registros Especiales"][
                            "egresos_urgencias_observacion"
                        ] = max(
                            mes_data["Registros Especiales"][
                                "egresos_urgencias_observacion"
                            ],
                            total_egreso,
                        )
                    elif cve_especialidad == "5005":  # Urgencias Observación Pediatría
                        mes_data["Registros Especiales"][
                            "urg_observacion_pediatria"
                        ] = max(
                            mes_data["Registros Especiales"][
                                "urg_observacion_pediatria"
                            ],
                            total_egreso,
                        )
                    elif cve_especialidad == "A600":  # Urgencias Tococirugía
                        mes_data["Registros Especiales"]["urgencias_tococirugia"] = max(
                            mes_data["Registros Especiales"]["urgencias_tococirugia"],
                            total_egreso,
                        )
                    elif cve_especialidad == "ECAM":  # ECAM
                        mes_data["Registros Especiales"]["ecam"] = max(
                            mes_data["Registros Especiales"]["ecam"], total_egreso
                        )

                filas_procesadas += 1

            except Exception as e:
                continue

        workbook.close()

        print(f"    Procesadas {filas_procesadas} filas de hospitalización")
        print(f"    Hospitales con datos: {len(datos_hospitalizacion)}")

        return datos_hospitalizacion

    except Exception as e:
        print(f" Error procesando Excel PARTEII: {e}")
        return {}


def procesar_hospitalizacion_hgz(clave_presupuestal, datos_hospitalizacion):
    """
    Procesa los datos de hospitalización para un hospital específico
    Genera la estructura mensual completa con todas las divisiones
    """
    # Meses en orden
    meses_orden = [
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    ]

    # Obtener datos de este hospital
    datos_hospital = datos_hospitalizacion.get(clave_presupuestal, {})

    # Estructura final
    hospitalizacion = {}

    for mes in meses_orden:
        mes_data = datos_hospital.get(
            mes,
            {
                "Cirugía": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                "Gineco-Obstetricia": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                "Medicina": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                "Pediatría": {"egresos": 0, "dias_paciente": 0, "camas": 0},
                "Registros Especiales": {
                    "cirugias_ambulatorias": 0,
                    "egresos_urgencias_observacion": 0,
                    "uci": 0,
                    "uci_neonato": 0,
                    "urg_reanimacion": 0,
                    "urg_observacion_intermedia": 0,
                    "urg_observacion_pediatria": 0,
                    "urgencias_tococirugia": 0,
                    "ecam": 0,
                },
            },
        )

        # Calcular totales
        total_egresos = (
            mes_data["Cirugía"]["egresos"]
            + mes_data["Gineco-Obstetricia"]["egresos"]
            + mes_data["Medicina"]["egresos"]
            + mes_data["Pediatría"]["egresos"]
        )

        total_dias_paciente = (
            mes_data["Cirugía"]["dias_paciente"]
            + mes_data["Gineco-Obstetricia"]["dias_paciente"]
            + mes_data["Medicina"]["dias_paciente"]
            + mes_data["Pediatría"]["dias_paciente"]
        )

        total_dias_cama = (
            mes_data["Cirugía"]["camas"]
            + mes_data["Gineco-Obstetricia"]["camas"]
            + mes_data["Medicina"]["camas"]
            + mes_data["Pediatría"]["camas"]
        )

        # Calcular indicadores
        porc_ocupacion = (
            round((total_dias_paciente / total_dias_cama * 100), 2)
            if total_dias_cama > 0
            else 0
        )
        prom_estancia = (
            round((total_dias_paciente / total_egresos), 2) if total_egresos > 0 else 0
        )
        ind_rotacion = (
            round((total_egresos / total_dias_cama), 2) if total_dias_cama > 0 else 0
        )

        hospitalizacion[mes] = {
            "egresos": {
                "total": total_egresos,
                "cirugia": mes_data["Cirugía"]["egresos"],
                "gineco_obstetricia": mes_data["Gineco-Obstetricia"]["egresos"],
                "medicina": mes_data["Medicina"]["egresos"],
                "pediatria": mes_data["Pediatría"]["egresos"],
            },
            "dias_paciente": {
                "total": total_dias_paciente,
                "cirugia": mes_data["Cirugía"]["dias_paciente"],
                "gineco_obstetricia": mes_data["Gineco-Obstetricia"]["dias_paciente"],
                "medicina": mes_data["Medicina"]["dias_paciente"],
                "pediatria": mes_data["Pediatría"]["dias_paciente"],
            },
            "dias_cama": {
                "total": total_dias_cama,
                "cirugia": mes_data["Cirugía"]["camas"],
                "gineco_obstetricia": mes_data["Gineco-Obstetricia"]["camas"],
                "medicina": mes_data["Medicina"]["camas"],
                "pediatria": mes_data["Pediatría"]["camas"],
            },
            "porcentaje_ocupacion": {
                "total": porc_ocupacion,
                "cirugia": round(
                    (
                        mes_data["Cirugía"]["dias_paciente"]
                        / mes_data["Cirugía"]["camas"]
                        * 100
                    ),
                    2,
                )
                if mes_data["Cirugía"]["camas"] > 0
                else 0,
                "gineco_obstetricia": round(
                    (
                        mes_data["Gineco-Obstetricia"]["dias_paciente"]
                        / mes_data["Gineco-Obstetricia"]["camas"]
                        * 100
                    ),
                    2,
                )
                if mes_data["Gineco-Obstetricia"]["camas"] > 0
                else 0,
                "medicina": round(
                    (
                        mes_data["Medicina"]["dias_paciente"]
                        / mes_data["Medicina"]["camas"]
                        * 100
                    ),
                    2,
                )
                if mes_data["Medicina"]["camas"] > 0
                else 0,
                "pediatria": round(
                    (
                        mes_data["Pediatría"]["dias_paciente"]
                        / mes_data["Pediatría"]["camas"]
                        * 100
                    ),
                    2,
                )
                if mes_data["Pediatría"]["camas"] > 0
                else 0,
            },
            "promedio_estancia": {
                "total": prom_estancia,
                "cirugia": round(
                    (
                        mes_data["Cirugía"]["dias_paciente"]
                        / mes_data["Cirugía"]["egresos"]
                    ),
                    2,
                )
                if mes_data["Cirugía"]["egresos"] > 0
                else 0,
                "gineco_obstetricia": round(
                    (
                        mes_data["Gineco-Obstetricia"]["dias_paciente"]
                        / mes_data["Gineco-Obstetricia"]["egresos"]
                    ),
                    2,
                )
                if mes_data["Gineco-Obstetricia"]["egresos"] > 0
                else 0,
                "medicina": round(
                    (
                        mes_data["Medicina"]["dias_paciente"]
                        / mes_data["Medicina"]["egresos"]
                    ),
                    2,
                )
                if mes_data["Medicina"]["egresos"] > 0
                else 0,
                "pediatria": round(
                    (
                        mes_data["Pediatría"]["dias_paciente"]
                        / mes_data["Pediatría"]["egresos"]
                    ),
                    2,
                )
                if mes_data["Pediatría"]["egresos"] > 0
                else 0,
            },
            "indice_rotacion": {
                "total": ind_rotacion,
                "cirugia": round(
                    (mes_data["Cirugía"]["egresos"] / mes_data["Cirugía"]["camas"]), 2
                )
                if mes_data["Cirugía"]["camas"] > 0
                else 0,
                "gineco_obstetricia": round(
                    (
                        mes_data["Gineco-Obstetricia"]["egresos"]
                        / mes_data["Gineco-Obstetricia"]["camas"]
                    ),
                    2,
                )
                if mes_data["Gineco-Obstetricia"]["camas"] > 0
                else 0,
                "medicina": round(
                    (mes_data["Medicina"]["egresos"] / mes_data["Medicina"]["camas"]), 2
                )
                if mes_data["Medicina"]["camas"] > 0
                else 0,
                "pediatria": round(
                    (mes_data["Pediatría"]["egresos"] / mes_data["Pediatría"]["camas"]),
                    2,
                )
                if mes_data["Pediatría"]["camas"] > 0
                else 0,
            },
            "registros_especiales": mes_data["Registros Especiales"],
        }

    return hospitalizacion


def procesar_productividad_servicios_medicos_hgz(
    clave_presupuestal, datos_consulta_externa
):
    """
    Procesa los datos de productividad en servicios médicos para hospitales HGZ
    Retorna estructura mensual con todos los indicadores
    """
    # Meses en orden
    meses_orden = [
        "ENERO",
        "FEBRERO",
        "MARZO",
        "ABRIL",
        "MAYO",
        "JUNIO",
        "JULIO",
        "AGOSTO",
        "SEPTIEMBRE",
        "OCTUBRE",
        "NOVIEMBRE",
        "DICIEMBRE",
    ]

    # Obtener datos de este hospital
    datos_hospital = datos_consulta_externa.get(clave_presupuestal, {})

    # Estructura final
    productividad_servicios_medicos = {}

    for mes in meses_orden:
        mes_data = datos_hospital.get(
            mes,
            {
                "total_consultas": 0,
                "consultas_especialidades": 0,
                "consultas_primera_vez": 0,
                "consultas_subsecuentes": 0,
                "consultas_urgencias": 0,
                "consultas_tococirrugia": 0,
                "servicios_profesionales": 0,
            },
        )

        productividad_servicios_medicos[mes] = {
            "total_consultas_unidad": mes_data["total_consultas"],
            "consultas_especialidades": mes_data["consultas_especialidades"],
            "consultas_primera_vez": mes_data["consultas_primera_vez"],
            "consultas_subsecuentes": mes_data["consultas_subsecuentes"],
            "consultas_urgencias": mes_data["consultas_urgencias"],
            "consultas_tococirrugia": mes_data["consultas_tococirrugia"],
            "servicios_profesionales": mes_data["servicios_profesionales"],
        }

    return productividad_servicios_medicos


def consultar_motivos_atencion_umf(clave_presupuestal):
    """
    Consulta los motivos de atención principales para una UMF específica
    Busca en todos los servidores SIAIS
    Servicio 04: Medicina Familiar
    Servicio 5A: Urgencias
    """
    # Calcular fechas del año en juego (26/12/año_anterior al 25/12/año_actual)
    año_actual = date.today().year
    # Formato YYYYMMDD (sin separadores) - el más compatible con SQL Server
    fecha_inicio = f"{año_actual - 1}1226"
    fecha_fin = f"{año_actual}1225"

    # Para mostrar al usuario
    fecha_inicio_display = f"26/12/{año_actual - 1}"
    fecha_fin_display = f"25/12/{año_actual}"

    servidores = get_siais_servers()

    print(
        f"    Consultando motivos en {len(servidores)} servidores ({fecha_inicio_display} al {fecha_fin_display})..."
    )

    query = """
    SELECT
        tbalteracion.Servicio,
        tbalteracion.Clave,
        cie10.DescripcionM,
        COUNT(*) AS Total
    FROM tbalteracion
    LEFT JOIN tb_paCIE10_Busqueda_Basica cie10
        ON cie10.Clave = tbalteracion.Clave
    WHERE
        tbalteracion.Servicio IN ('04','5A')
        AND tbalteracion.Fecha >= %s
        AND tbalteracion.Fecha <= %s
        AND tbalteracion.DiagPrin = 1
        AND tbalteracion.CvePresup = %s
    GROUP BY
        tbalteracion.Servicio,
        tbalteracion.Clave,
        cie10.DescripcionM
    ORDER BY
        tbalteracion.Servicio, Total DESC
    """

    # Diccionarios para acumular resultados de todos los servidores
    motivos_acumulados = {}

    # DEBUG
    servidores_exitosos = 0
    servidores_con_datos = 0
    total_registros = 0
    errores_conexion = []

    # Buscar en cada servidor
    if not SIAIS_DB_CONFIG["user"] or not SIAIS_DB_CONFIG["password"]:
        print("      Variables FT_SIAIS_DB_USER y FT_SIAIS_DB_PASSWORD no configuradas")
        return {"medicina_familiar": [], "urgencias": [], "periodo": "No disponible"}

    for idx, servidor_ip in enumerate(servidores, 1):
        try:
            print(
                f"         Intentando servidor {idx}/{len(servidores)}: {servidor_ip}...",
                end="",
                flush=True,
            )
            conn = pymssql.connect(
                server=f"{servidor_ip}:{SIAIS_DB_CONFIG['port']}",
                user=SIAIS_DB_CONFIG["user"],
                password=SIAIS_DB_CONFIG["password"],
                database=SIAIS_DB_CONFIG["database"],
                timeout=10,
                login_timeout=5,
            )
            print("  Conectado")

            cursor = conn.cursor(as_dict=True)
            cursor.execute(query, (fecha_inicio, fecha_fin, clave_presupuestal))
            resultados = cursor.fetchall()

            # DEBUG
            servidores_exitosos += 1
            if len(resultados) > 0:
                servidores_con_datos += 1
                total_registros += len(resultados)
                print(
                    f"          {servidor_ip}: {len(resultados)} registros | CvePresup: {clave_presupuestal}"
                )

            # Acumular resultados por servicio y clave CIE10
            for row in resultados:
                servicio = row["Servicio"].strip()
                clave = row["Clave"]
                descripcion = row["DescripcionM"] or "Sin descripción"
                total = row["Total"]

                key = f"{servicio}_{clave}"

                if key not in motivos_acumulados:
                    motivos_acumulados[key] = {
                        "servicio": servicio,
                        "clave": clave,
                        "descripcion": descripcion,
                        "total": 0,
                    }

                motivos_acumulados[key]["total"] += total

            cursor.close()
            conn.close()

        except Exception as e:
            # DEBUG: Capturar error
            print(f"  Error")
            errores_conexion.append(f"{servidor_ip}: {str(e)}")
            continue

    # DEBUG: Mostrar resumen
    print(
        f"          Servidores consultados: {servidores_exitosos}/{len(servidores)}"
    )
    print(f"          Servidores con datos: {servidores_con_datos}")
    print(f"          Total registros: {total_registros}")

    # Mostrar errores si no se pudo consultar todos los servidores
    if errores_conexion:
        servidores_fallidos = len(errores_conexion)
        if servidores_fallidos > 0:
            print(f"           Servidores con error: {servidores_fallidos}")
            # Mostrar solo los primeros 3 errores para no saturar
            for i, error in enumerate(errores_conexion[:3]):
                print(f"            - {error.split(':')[0]}")
            if servidores_fallidos > 3:
                print(f"            - ... y {servidores_fallidos - 3} más")

    # Separar por servicio y ordenar por total
    motivos_04 = []
    motivos_5A = []

    for key, motivo in motivos_acumulados.items():
        servicio = motivo["servicio"]
        item = {
            "clave": motivo["clave"],
            "descripcion": motivo["descripcion"],
            "total": motivo["total"],
        }

        if servicio == "04":
            motivos_04.append(item)
        elif servicio == "5A":
            motivos_5A.append(item)

    # Ordenar por total descendente y tomar top 25
    motivos_04 = sorted(motivos_04, key=lambda x: x["total"], reverse=True)[:25]
    motivos_5A = sorted(motivos_5A, key=lambda x: x["total"], reverse=True)[:25]

    print(f"       Encontrados: {len(motivos_04)} MF, {len(motivos_5A)} URG")

    return {
        "medicina_familiar": motivos_04,
        "urgencias": motivos_5A,
        "periodo": f"{fecha_inicio_display} al {fecha_fin_display}",
    }


def consultar_motivos_atencion_hospital(clave_presupuestal):
    """
    Consulta los motivos de atención principales para un Hospital (HGZ, HGSZMF)
    Busca en todos los servidores SIAIS
    Servicios: Especialidades (excluyendo 04) y 5A (Urgencias)
    """
    año_actual = date.today().year
    fecha_inicio = f"{año_actual - 1}1226"
    fecha_fin = f"{año_actual}1225"

    fecha_inicio_display = f"26/12/{año_actual - 1}"
    fecha_fin_display = f"25/12/{año_actual}"

    servidores = get_siais_servers()

    print(
        f"    Consultando motivos hospitalarios en {len(servidores)} servidores ({fecha_inicio_display} al {fecha_fin_display})..."
    )

    # Query diferente: busca especialidades (NOT IN '04') y urgencias '5A'
    query = """
    SELECT
        tbalteracion.Servicio,
        tbalteracion.Clave,
        cie10.DescripcionM,
        COUNT(*) AS Total
    FROM tbalteracion
    LEFT JOIN tb_paCIE10_Busqueda_Basica cie10
        ON cie10.Clave = tbalteracion.Clave
    WHERE
        (tbalteracion.Servicio NOT IN ('04') OR tbalteracion.Servicio = '5A')
        AND tbalteracion.Fecha >= %s
        AND tbalteracion.Fecha <= %s
        AND tbalteracion.DiagPrin = 1
        AND tbalteracion.CvePresup = %s
    GROUP BY
        tbalteracion.Servicio,
        tbalteracion.Clave,
        cie10.DescripcionM
    ORDER BY
        tbalteracion.Servicio, Total DESC
    """

    motivos_acumulados = {}
    servidores_exitosos = 0
    servidores_con_datos = 0
    total_registros = 0
    errores_conexion = []

    if not SIAIS_DB_CONFIG["user"] or not SIAIS_DB_CONFIG["password"]:
        print("      Variables FT_SIAIS_DB_USER y FT_SIAIS_DB_PASSWORD no configuradas")
        return {"especialidades": [], "urgencias": [], "periodo": "No disponible"}

    for idx, servidor_ip in enumerate(servidores, 1):
        try:
            print(
                f"         Intentando servidor {idx}/{len(servidores)}: {servidor_ip}...",
                end="",
                flush=True,
            )
            conn = pymssql.connect(
                server=f"{servidor_ip}:{SIAIS_DB_CONFIG['port']}",
                user=SIAIS_DB_CONFIG["user"],
                password=SIAIS_DB_CONFIG["password"],
                database=SIAIS_DB_CONFIG["database"],
                timeout=10,
                login_timeout=5,
            )
            print("  Conectado")

            cursor = conn.cursor(as_dict=True)
            cursor.execute(query, (fecha_inicio, fecha_fin, clave_presupuestal))
            resultados = cursor.fetchall()

            servidores_exitosos += 1
            if len(resultados) > 0:
                servidores_con_datos += 1
                total_registros += len(resultados)
                print(
                    f"          {servidor_ip}: {len(resultados)} registros | CvePresup: {clave_presupuestal}"
                )

            for row in resultados:
                servicio = row["Servicio"].strip()
                clave = row["Clave"]
                descripcion = row["DescripcionM"] or "Sin descripción"
                total = row["Total"]

                key = f"{servicio}_{clave}"

                if key not in motivos_acumulados:
                    motivos_acumulados[key] = {
                        "servicio": servicio,
                        "clave": clave,
                        "descripcion": descripcion,
                        "total": 0,
                    }

                motivos_acumulados[key]["total"] += total

            cursor.close()
            conn.close()

        except Exception as e:
            print(f"  Error")
            errores_conexion.append(f"{servidor_ip}: {str(e)}")
            continue

    print(
        f"          Servidores consultados: {servidores_exitosos}/{len(servidores)}"
    )
    print(f"          Servidores con datos: {servidores_con_datos}")
    print(f"          Total registros: {total_registros}")

    if errores_conexion:
        servidores_fallidos = len(errores_conexion)
        if servidores_fallidos > 0:
            print(f"           Servidores con error: {servidores_fallidos}")

    # Separar especialidades (todos menos 04 y 5A) y urgencias (5A)
    motivos_especialidades = []
    motivos_urgencias = []

    for key, motivo in motivos_acumulados.items():
        servicio = motivo["servicio"]

        motivo_formateado = {
            "clave": motivo["clave"],
            "descripcion": motivo["descripcion"],
            "total": motivo["total"],
        }

        if servicio == "5A":
            motivos_urgencias.append(motivo_formateado)
        elif servicio != "04":  # Cualquier otro servicio que no sea 04
            motivos_especialidades.append(motivo_formateado)

    # Ordenar y limitar a top 25
    motivos_especialidades = sorted(
        motivos_especialidades, key=lambda x: x["total"], reverse=True
    )[:25]
    motivos_urgencias = sorted(
        motivos_urgencias, key=lambda x: x["total"], reverse=True
    )[:25]

    return {
        "especialidades": motivos_especialidades,
        "urgencias": motivos_urgencias,
        "periodo": f"{fecha_inicio_display} al {fecha_fin_display}",
    }


def crear_mapeo_recursos():
    """Crea el mapeo entre códigos del IFU y nombres de recursos"""
    return {
        # CONSULTORIOS
        "70000": "total_consultorio_unidad",
        "70101": "medicina_familiar",
        "71200": "especialidad",
        "70900": "urgencias_medicas",
        "71000": "atn_medica_continua",
        "70105": "cadimss",
        "70106": "planificacion_familiar",
        "70200": "estomatologia",
        "70800": "nutricion_dietetica",
        "70150": "profesionales",
        # CAMAS CENSABLES
        "50100": "total_censables",
        "56000": "medicina_interna",
        "55000": "cirugia",
        "51200": "gineco_obstetricia",
        "53620": "pediatria",
        "53600": "cuna_cirugia_pediatrica",
        # CAMAS NO CENSABLES
        "60000": "no_censables",
        "60300": "camillas_amc",
        "60301": "observacion_adulto",
        "60302": "observacion_pediatrica",
        # QUIRÓFANOS
        "80364": "quirofanos",
        "80403": "sala_expulsion",
        "80402": "mixta_cirugia_tococirugia",
        "80361": "urgencias_quirofano",
        "80407": "tococirugia",
        "80354": "hibrida",
        # URGENCIAS
        "60100": "total_camas_urgencias",
        "60202": "cama_observ_adulto",
        "60201": "cama_observ_pediatrica",
        "60104": "cama_corta_estancia",
        "60105": "reanimacion_adulto",
        # EQUIPOS DIAGNÓSTICO
        "81402": "peine_laboratorio_clinico",
        "80602": "tomografos",
        "80658": "rayos_x_fijos",
        "80659": "rayos_x_transportable",
        "80661": "mastografo_digital",
        "80751": "electrocardiografo",
        "80753": "electromiografo",
        "80752": "electroencefalografo",
        "80754": "ecocardiografo",
        # PROCEDIMIENTOS
        "80408": "procedimientos_quirurgicos",
        "80451": "endoscopia_altas",
        "80452": "endoscopia_bajas",
        "80651": "radiodiagnostico",
        # HEMODIÁLISIS
        "81301": "maquinas_hemodialisis",
        "81303": "areas_hemodialisis",
        # AMBULANCIAS
        "81900": "ambulancias",
        "81901": "ambulancias_equipadas",
        # SERVICIOS ESPECIALES
        "80151": "prevenimss",
        "70104": "enf_esp_med_familiar",
        "80101": "atencion_orientacion_dh",
        "77050": "banco_sangre",
        "70107": "aux_medicina_familiar",
        "70400": "consultorio_enf_prenatal",
        "70500": "consultorio_enf_cronicos",
        "70600": "consultorio_enf_materno_infantil",
        "70700": "consultorio_optometria",
        "75400": "consultorio_psicologia",
    }


def integrar_recursos_en_fichas(
    fichas_originales,
    datos_ifu,
    datos_productividad=None,
    datos_consulta_externa_hgz=None,
    datos_hospitalizacion_hgz=None,
    datos_egresos_hospital=None,
    datos_intervenciones_hospital=None,
    catalogo_cie10=None,
    catalogo_cie9=None,
    archivo_consulta_externa_hgz=None,
    procesar_solo_hgz=False,
):
    """Integra los datos de recursos del IFU y productividad en las fichas"""

    mapeo_recursos = crear_mapeo_recursos()

    # Crear un índice de unidades del IFU por denominación
    indice_ifu = {}
    if not procesar_solo_hgz:
        for unidad_ifu in datos_ifu:
            denominacion = unidad_ifu.get("denominacion_completa", "")
            if denominacion:
                indice_ifu[denominacion] = unidad_ifu.get("datos_ifu", {})

        print(f" Índice IFU creado con {len(indice_ifu)} unidades")
    else:
        print(f"SKIP:  Modo HGZ: Omitiendo creación de índice IFU")

    # Si no se proveen datos de productividad, usar diccionario vacío
    if datos_productividad is None:
        datos_productividad = {}

    # Si no se proveen datos de consulta externa HGZ, usar diccionario vacío
    if datos_consulta_externa_hgz is None:
        datos_consulta_externa_hgz = {}

    # Si no se proveen datos de hospitalización HGZ, usar diccionario vacío
    if datos_hospitalizacion_hgz is None:
        datos_hospitalizacion_hgz = {}

    # Si no se proveen datos de motivos hospitalarios, usar diccionarios vacíos
    if datos_egresos_hospital is None:
        datos_egresos_hospital = {}
    if datos_intervenciones_hospital is None:
        datos_intervenciones_hospital = {}
    if catalogo_cie10 is None:
        catalogo_cie10 = {}
    if catalogo_cie9 is None:
        catalogo_cie9 = {}

    unidades_actualizadas = 0
    unidades_sin_recursos = 0
    unidades_con_productividad = 0
    hospitales_con_productividad_servicios = 0
    hospitales_con_hospitalizacion = 0
    hospitales_con_motivos = 0

    # Procesar cada ficha
    for unidad in fichas_originales.get("unidades", []):
        info_general = unidad.get("informacion_general", {})
        denominacion = info_general.get("denominacion_completa", "")
        nombre_unidad = info_general.get("nombre_unidad", "Sin nombre")
        tipo_unidad = info_general.get("tipo_unidad", "")

        # Intentar obtener la clave presupuestal de datos_administrativos_completos primero
        datos_admin = info_general.get("datos_administrativos_completos", {})
        clave_presupuestal = datos_admin.get(
            "Clave Presupuestal", info_general.get("clave_presupuestal", "")
        )

        # Convertir a string si es numérico
        if isinstance(clave_presupuestal, (int, float)):
            clave_presupuestal = str(int(clave_presupuestal))

        # Si estamos en modo HGZ, saltar todo excepto hospitales
        if procesar_solo_hgz and tipo_unidad not in [
            "Hospital General de Zona",
            "Hospital General de Subzona",
            "Hospital General de Subzona con Medicina Familiar",
        ]:
            continue

        # RECURSOS IFU (solo si no es modo HGZ)
        if not procesar_solo_hgz:
            if denominacion in indice_ifu:
                # Obtener datos de recursos del IFU
                datos_recursos_ifu = indice_ifu[denominacion]

                # Crear estructura de recursos
                recursos = {
                    "consultorios": {},
                    "camas_censables": {},
                    "camas_no_censables": {},
                    "quirofanos": {},
                    "urgencias": {},
                    "equipos_diagnostico": {},
                    "procedimientos": {},
                    "hemodialisis": {},
                    "ambulancias": {},
                    "servicios_especiales": {},
                }

                # Mapear cada código a su categoría y nombre
                for codigo, valor in datos_recursos_ifu.items():
                    if codigo in mapeo_recursos:
                        nombre_recurso = mapeo_recursos[codigo]
                        valor_final = valor if valor is not None else 0

                        # Asignar a la categoría correspondiente
                        if codigo in [
                            "70000",
                            "70101",
                            "71200",
                            "70900",
                            "71000",
                            "70105",
                            "70106",
                            "70200",
                            "70800",
                            "70150",
                            "70107",
                            "70400",
                            "70500",
                            "70600",
                            "70700",
                            "75400",
                        ]:
                            recursos["consultorios"][nombre_recurso] = valor_final
                        elif codigo in [
                            "50100",
                            "56000",
                            "55000",
                            "51200",
                            "53620",
                            "53600",
                        ]:
                            recursos["camas_censables"][nombre_recurso] = valor_final
                        elif codigo in ["60000", "60300", "60301", "60302"]:
                            recursos["camas_no_censables"][nombre_recurso] = valor_final
                        elif codigo in [
                            "80364",
                            "80403",
                            "80402",
                            "80361",
                            "80407",
                            "80354",
                        ]:
                            recursos["quirofanos"][nombre_recurso] = valor_final
                        elif codigo in ["60100", "60202", "60201", "60104", "60105"]:
                            recursos["urgencias"][nombre_recurso] = valor_final
                        elif codigo in [
                            "81402",
                            "80602",
                            "80658",
                            "80659",
                            "80661",
                            "80751",
                            "80753",
                            "80752",
                            "80754",
                        ]:
                            recursos["equipos_diagnostico"][nombre_recurso] = (
                                valor_final
                            )
                        elif codigo in ["80408", "80451", "80452", "80651"]:
                            recursos["procedimientos"][nombre_recurso] = valor_final
                        elif codigo in ["81301", "81303"]:
                            recursos["hemodialisis"][nombre_recurso] = valor_final
                        elif codigo in ["81900", "81901"]:
                            recursos["ambulancias"][nombre_recurso] = valor_final
                        elif codigo in ["80151", "70104", "80101", "77050"]:
                            recursos["servicios_especiales"][nombre_recurso] = (
                                valor_final
                            )

                # Agregar recursos a la unidad
                unidad["recursos_para_la_salud"] = recursos
                unidades_actualizadas += 1

                print(f"    {nombre_unidad}: Recursos integrados")

            else:
                unidades_sin_recursos += 1
                print(f"    {nombre_unidad}: Sin datos de recursos")

        # PRODUCTIVIDAD Y MOTIVOS UMF (solo si no es modo HGZ)
        if not procesar_solo_hgz and tipo_unidad == "Unidad de Medicina Familiar":
            # Agregar productividad
            if (
                datos_productividad
                and clave_presupuestal
                and clave_presupuestal in datos_productividad
            ):
                productividad = procesar_productividad_unidad(
                    clave_presupuestal, datos_productividad
                )
                unidad["productividad_servicios_auxiliares"] = productividad
                unidades_con_productividad += 1
                print(f"    {nombre_unidad}: Productividad integrada")

            # Agregar motivos de atención
            if clave_presupuestal:
                print(
                    f"    {nombre_unidad}: Consultando motivos de atención en servidores SIAIS..."
                )
                try:
                    motivos = consultar_motivos_atencion_umf(clave_presupuestal)
                    unidad["motivos_atencion"] = motivos
                    print(
                        f"    {nombre_unidad}: Motivos de atención integrados ({len(motivos['medicina_familiar'])} MF, {len(motivos['urgencias'])} URG)"
                    )
                except Exception as e:
                    print(
                        f"     {nombre_unidad}: Error al consultar motivos de atención: {str(e)}"
                    )
                    unidad["motivos_atencion"] = {
                        "medicina_familiar": [],
                        "urgencias": [],
                    }

        # PRODUCTIVIDAD SERVICIOS MEDICOS HGZ (siempre para hospitales)
        if tipo_unidad in [
            "Hospital General de Zona",
            "Hospital General de Subzona",
            "Hospital General de Subzona con Medicina Familiar",
        ]:
            print(f"    {nombre_unidad} - Tipo: {tipo_unidad}")
            print(
                f"      Clave Presupuestal: '{clave_presupuestal}' (tipo: {type(clave_presupuestal).__name__})"
            )
            print(
                f"      Tiene datos_consulta_externa_hgz: {bool(datos_consulta_externa_hgz)}"
            )
            print(
                f"      Clave en datos_consulta_externa_hgz: {clave_presupuestal in datos_consulta_externa_hgz if datos_consulta_externa_hgz else False}"
            )
            if datos_hospitalizacion_hgz:
                print(
                    f"      Clave en datos_hospitalizacion_hgz: {clave_presupuestal in datos_hospitalizacion_hgz}"
                )

            if (
                datos_consulta_externa_hgz
                and clave_presupuestal
                and clave_presupuestal in datos_consulta_externa_hgz
            ):
                productividad_servicios = procesar_productividad_servicios_medicos_hgz(
                    clave_presupuestal, datos_consulta_externa_hgz
                )
                unidad["productividad_servicios_medicos"] = productividad_servicios
                hospitales_con_productividad_servicios += 1
                print(
                    f"    {nombre_unidad}: Productividad Servicios Médicos integrada"
                )
            else:
                print(f"     {nombre_unidad}: No se pudo integrar productividad")

            # DATOS DE HOSPITALIZACIÓN (PARTEII)
            if (
                datos_hospitalizacion_hgz
                and clave_presupuestal
                and clave_presupuestal in datos_hospitalizacion_hgz
            ):
                hospitalizacion = procesar_hospitalizacion_hgz(
                    clave_presupuestal, datos_hospitalizacion_hgz
                )
                unidad["hospitalizacion"] = hospitalizacion
                hospitales_con_hospitalizacion += 1
                print(f"    {nombre_unidad}: Datos de Hospitalización integrados")
            else:
                print(
                    f"     {nombre_unidad}: No se encontraron datos de hospitalización"
                )

            # PRODUCTIVIDAD SERVICIOS AUXILIARES (para hospitales también)
            if (
                datos_productividad
                and clave_presupuestal
                and clave_presupuestal in datos_productividad
            ):
                productividad = procesar_productividad_unidad(
                    clave_presupuestal, datos_productividad
                )
                unidad["productividad_servicios_auxiliares"] = productividad
                print(
                    f"    {nombre_unidad}: Productividad Servicios Auxiliares integrada"
                )

            # MOTIVOS DE ATENCIÓN HOSPITALARIA (diagnósticos de egreso + procedimientos quirúrgicos)
            if datos_egresos_hospital or datos_intervenciones_hospital:
                if clave_presupuestal:
                    motivos_hospital = procesar_motivos_hospital(
                        clave_presupuestal,
                        datos_egresos_hospital,
                        datos_intervenciones_hospital,
                        catalogo_cie10,
                        catalogo_cie9,
                    )
                    # Solo agregar si tiene al menos algún dato
                    if (
                        motivos_hospital["diagnosticos_egreso"]
                        or motivos_hospital["procedimientos_quirurgicos"]
                    ):
                        unidad["motivos_atencion_hospital"] = motivos_hospital
                        hospitales_con_motivos += 1
                        total_diag = len(motivos_hospital["diagnosticos_egreso"])
                        total_proc = len(motivos_hospital["procedimientos_quirurgicos"])
                        print(
                            f"    {nombre_unidad}: Motivos hospitalarios integrados ({total_diag} diagnósticos, {total_proc} procedimientos)"
                        )

                        # AGREGAR MOTIVOS DE CONSULTA EXTERNA (Especialidades + Urgencias)
                        print(
                            f"\n    Procesando motivos de Consulta Externa para {nombre_unidad}..."
                        )
                        archivo_consulta_externa = archivo_consulta_externa_hgz or str(
                            DATA_INPUT_DIR / "Consulta_Externa_Diaria_2025.xlsx"
                        )

                        motivos_consulta = procesar_consulta_externa_motivos_hospital(
                            archivo_consulta_externa, clave_presupuestal, catalogo_cie10
                        )

                        if motivos_consulta:
                            # Agregar al campo motivos_atencion_hospital existente
                            unidad["motivos_atencion_hospital"][
                                "especialidades_consulta"
                            ] = motivos_consulta["especialidades"]
                            unidad["motivos_atencion_hospital"][
                                "urgencias_consulta"
                            ] = motivos_consulta["urgencias"]

                            total_esp = len(motivos_consulta["especialidades"])
                            total_urg = len(motivos_consulta["urgencias"])
                            print(
                                f"    {nombre_unidad}: Motivos consulta externa integrados ({total_esp} especialidades, {total_urg} urgencias)"
                            )

            # MOTIVOS DE MEDICINA FAMILIAR (SIAIS) - Solo para HGSZMF
            if tipo_unidad == "Hospital General de Subzona con Medicina Familiar":
                if not procesar_solo_hgz:
                    print(
                        f"\n    {nombre_unidad}: Consultando motivos de Medicina Familiar en servidores SIAIS..."
                    )
                    try:
                        motivos_mf = consultar_motivos_atencion_umf(clave_presupuestal)
                        if motivos_mf:
                            # Agregar al campo motivos_atencion para mantener compatibilidad con UMF
                            unidad["motivos_atencion"] = motivos_mf
                            print(
                                f"    {nombre_unidad}: Motivos Medicina Familiar integrados ({len(motivos_mf['medicina_familiar'])} MF, {len(motivos_mf['urgencias'])} URG)"
                            )
                    except Exception as e:
                        print(
                            f"     Error consultando Medicina Familiar para {nombre_unidad}: {e}"
                        )

            # NOTA: Consulta SIAIS desactivada para hospitales HGZ (usuario no la necesita)
            # Si se necesita en el futuro, descomentar este bloque:
            # if tipo_unidad in ['Hospital General de Zona', 'Hospital General de Subzona']:
            #     if not procesar_solo_hgz:
            #         print(f"\n Consultando motivos SIAIS para {nombre_unidad}...")
            #         try:
            #             motivos_siais = consultar_motivos_atencion_hospital(clave_presupuestal)
            #             if motivos_siais:
            #                 unidad['motivos_atencion'] = motivos_siais
            #                 total_esp = len(motivos_siais.get('especialidades', []))
            #                 total_urg = len(motivos_siais.get('urgencias', []))
            #                 print(f"    {nombre_unidad}: Motivos SIAIS integrados ({total_esp} especialidades, {total_urg} urgencias)")
            #         except Exception as e:
            #             print(f"     Error consultando SIAIS para {nombre_unidad}: {e}")

    print(f"\n Resumen de integración:")
    print(f"    Unidades con recursos: {unidades_actualizadas}")
    print(f"    Unidades sin recursos: {unidades_sin_recursos}")
    print(
        f"    UMF con productividad servicios auxiliares: {unidades_con_productividad}"
    )
    print(
        f"    Hospitales con productividad servicios médicos: {hospitales_con_productividad_servicios}"
    )
    print(
        f"    Hospitales con datos de hospitalización: {hospitales_con_hospitalizacion}"
    )
    print(f"    Hospitales con motivos de atención: {hospitales_con_motivos}")

    return fichas_originales


def guardar_fichas_actualizadas(fichas_actualizadas):
    """Guarda las fichas con los recursos integrados"""
    archivo_salida = str(DATA_OUTPUT_DIR / "fichas_completas_con_recursos.json")
    os.makedirs(DATA_OUTPUT_DIR, exist_ok=True)

    # Crear backup del archivo anterior si existe
    if os.path.exists(archivo_salida):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archivo_backup = str(
            DATA_OUTPUT_DIR / f"fichas_completas_con_recursos_backup_{timestamp}.json"
        )
        os.rename(archivo_salida, archivo_backup)
        print(f" Backup creado: {archivo_backup}")

    with open(archivo_salida, "w", encoding="utf-8") as f:
        json.dump(fichas_actualizadas, f, indent=2, ensure_ascii=False, default=str)

    tamaño = os.path.getsize(archivo_salida)
    print(f" Fichas actualizadas guardadas: {archivo_salida} ({tamaño:,} bytes)")

    return archivo_salida


def main():
    print(" INTEGRADOR DE RECURSOS IFU A FICHAS TÉCNICAS")
    print("=" * 60)

    # Preguntar al usuario qué quiere procesar
    print("\n¿Qué deseas procesar?")
    print(
        "  1) Solo Hospitales (HGZ) - Actualizar solo productividad servicios médicos"
    )
    print(
        "  2) Todo completo (UMF + HGZ) - Incluye motivos de atención, productividad UMF y HGZ"
    )

    opcion = input("\nSelecciona una opción (1 o 2): ").strip()

    procesar_solo_hgz = opcion == "1"

    if procesar_solo_hgz:
        print("\n Modo: SOLO HOSPITALES (HGZ)")
        print("   Se procesará únicamente la productividad en servicios médicos")
        print(
            "   No se consultarán servidores SIAIS ni se cargará productividad de UMF"
        )
    else:
        print("\n Modo: PROCESO COMPLETO (UMF + HGZ)")
        print(
            "   Se procesará todo: recursos, productividad UMF, motivos de atención y HGZ"
        )

    # 1. Cargar datos del IFU (solo si es proceso completo)
    datos_ifu = []
    if not procesar_solo_hgz:
        datos_ifu = cargar_datos_ifu()
        if not datos_ifu:
            return
    else:
        print("\nSKIP:  Omitiendo carga de datos IFU (no necesario para HGZ)")

    # 2. Cargar fichas originales
    try:
        fichas_originales = cargar_fichas_originales()
    except Exception as e:
        print(f" Error cargando fichas: {e}")
        return

    # 3. Cargar datos de productividad de servicios auxiliares (UMF + HGZ - siempre)
    print(f"\n Cargando datos de productividad de servicios auxiliares...")
    datos_productividad = cargar_productividad_csv()

    # 4. Cargar datos de consulta externa (HGZ)
    print(f"\n Cargando datos de Consulta Externa HGZ...")
    # Intentar encontrar el archivo en diferentes rutas posibles
    rutas_posibles_consulta = [
        str(DATA_INPUT_DIR / "Consulta_Externa_Diaria_2025.xlsx"),
        str(DATA_INPUT_DIR / "Consulta_Externa_Diaria_2025.xls"),
        str(BASE_DIR / "data_sources" / "Consulta_Externa_Diaria_2025.xlsx"),
        str(BASE_DIR / "data_sources" / "Consulta_Externa_Diaria_2025.xls"),
    ]

    datos_consulta_externa_hgz = {}
    archivo_consulta_externa_hgz = None
    for ruta in rutas_posibles_consulta:
        if os.path.exists(ruta):
            archivo_consulta_externa_hgz = ruta
            datos_consulta_externa_hgz = cargar_consulta_externa_hgz(ruta)
            print(f"\n    Claves presupuestales encontradas en Excel:")
            for clave in datos_consulta_externa_hgz.keys():
                print(f"      - {clave}")
            break
    else:
        print(
            "  No se encontró el archivo de Consulta Externa HGZ en ninguna de las rutas esperadas"
        )
        print("   Buscando en:")
        for ruta in rutas_posibles_consulta:
            print(f"      - {ruta}")

    # 5. Cargar datos de hospitalización (PARTEII)
    print(f"\n Cargando datos de Hospitalización (PARTEII)...")
    rutas_posibles_parteii = [
        str(DATA_INPUT_DIR / "PARTEII_2025.xlsx"),
        str(BASE_DIR / "data_sources" / "PARTEII_2025.xlsx"),
    ]

    datos_hospitalizacion_hgz = {}
    for ruta in rutas_posibles_parteii:
        if os.path.exists(ruta):
            datos_hospitalizacion_hgz = cargar_hospitalizacion_parteii(ruta)
            print(f"\n    Claves presupuestales con datos de hospitalización:")
            for clave in datos_hospitalizacion_hgz.keys():
                print(f"      - {clave}")
            break
    else:
        print(
            "  No se encontró el archivo PARTEII_2025.xlsx en ninguna de las rutas esperadas"
        )
        print("   Buscando en:")
        for ruta in rutas_posibles_parteii:
            print(f"      - {ruta}")

    # 6. Cargar catálogo de descripciones CIE10/CIE9
    print(f"\n Cargando catálogo de descripciones...")
    catalogo_cie10, catalogo_cie9 = cargar_catalogo_descripciones()

    # 7. Cargar datos de egresos de pacientes
    print(f"\n Cargando datos de Egresos de Pacientes...")
    rutas_posibles_egresos = [
        str(DATA_INPUT_DIR / "Egresos_Pacientes_2025.xlsx"),
        str(BASE_DIR / "data_sources" / "Egresos_Pacientes_2025.xlsx"),
    ]

    datos_egresos_hospital = {}
    for ruta in rutas_posibles_egresos:
        if os.path.exists(ruta):
            datos_egresos_hospital = cargar_egresos_pacientes_hospital(ruta)
            break
    else:
        print("  No se encontró el archivo de Egresos de Pacientes")

    # 8. Cargar datos de intervenciones quirúrgicas
    print(f"\n Cargando datos de Intervenciones Quirúrgicas...")
    rutas_posibles_intervenciones = [
        str(DATA_INPUT_DIR / "Intervenciones_Quirurgicas_2025.xlsx"),
        str(BASE_DIR / "data_sources" / "Intervenciones_Quirurgicas_2025.xlsx"),
    ]

    datos_intervenciones_hospital = {}
    for ruta in rutas_posibles_intervenciones:
        if os.path.exists(ruta):
            datos_intervenciones_hospital = cargar_intervenciones_quirurgicas_hospital(
                ruta
            )
            break
    else:
        print("  No se encontró el archivo de Intervenciones Quirúrgicas")

    # 9. Integrar recursos y productividad
    print(f"\n Integrando recursos y productividad en fichas...")
    fichas_actualizadas = integrar_recursos_en_fichas(
        fichas_originales,
        datos_ifu,
        datos_productividad,
        datos_consulta_externa_hgz,
        datos_hospitalizacion_hgz,
        datos_egresos_hospital,
        datos_intervenciones_hospital,
        catalogo_cie10,
        catalogo_cie9,
        archivo_consulta_externa_hgz=archivo_consulta_externa_hgz,
        procesar_solo_hgz=procesar_solo_hgz,
    )

    # 5. Guardar fichas actualizadas
    archivo_guardado = guardar_fichas_actualizadas(fichas_actualizadas)

    print(f"\n Proceso completado exitosamente")
    print(f" Fichas con recursos: {archivo_guardado}")
    print(f" Ejemplo de uso: ejemplo_uso_recursos.js")
    print(f"\n Ahora puedes usar los recursos en tus fichas técnicas")


if __name__ == "__main__":
    main()
